import os
import logging
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

excel_lock = threading.Lock() 

# --- Local Excel File Path and Initialization ---
def get_local_excel_file_path(username, current_datetime, base_folder):
    """
    Generates the full path for the weekly Excel file for a given user.
    Format: base_folder/username/YYYY-WNN-username.xlsx
    """
    year = current_datetime.year
    week_number = current_datetime.isocalendar()[1]
    
    user_excel_folder = os.path.join(base_folder, username)
    os.makedirs(user_excel_folder, exist_ok=True) 

    excel_filename_weekly = f"{year}-W{week_number:02d}-{username}.xlsx"
    full_path = os.path.join(user_excel_folder, excel_filename_weekly)
    return full_path

def initialize_local_excel_file(username, current_datetime, base_folder):
    """
    Initializes the Excel file for a specific user and week, if it doesn't exist.
    """
    excel_file_path = get_local_excel_file_path(username, current_datetime, base_folder)
    sheet_name = "ImageMetadata"

    with excel_lock: # Use Lock to prevent concurrent file access
        if not os.path.exists(excel_file_path):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                ws.append(["ID (username)", "Bot Timestamp", "Image Log Name", "Extracted Image Timestamp"])
                wb.save(excel_file_path)
                logging.info(f"New local Excel file '{excel_file_path}' created with '{sheet_name}' sheet and headers.")
            except Exception as e:
                logging.error(f"Failed to create local Excel file '{excel_file_path}': {e}")
        else:
            try:
                wb = load_workbook(excel_file_path)
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    ws.append(["ID (username)", "Bot Timestamp", "Image Log Name", "Extracted Image Timestamp"])
                    logging.info(f"Created new sheet '{sheet_name}' in local Excel '{excel_file_path}' with headers.")
                else:
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in ws[1]]
                    if "Extracted Image Timestamp" not in headers:
                        ws.cell(row=1, column=len(headers) + 1, value="Extracted Image Timestamp")
                        logging.info(f"Added 'Extracted Image Timestamp' column to '{sheet_name}' in local Excel '{excel_file_path}'.")
                wb.save(excel_file_path)
            except Exception as e:
                logging.warning(f"Could not initialize sheet '{sheet_name}' in local Excel '{excel_file_path}': {e}")

# --- Append Data Functions ---
def append_to_local_excel(username, bot_timestamp, filename, extracted_image_timestamp_str, current_datetime, base_folder):
    """
    Appends image metadata to the correct weekly Excel file for the user.
    It will first ensure the Excel file exists and is initialized.
    """
    excel_file_path = get_local_excel_file_path(username, current_datetime, base_folder)
    
    initialize_local_excel_file(username, current_datetime, base_folder)

    with excel_lock:
        try:
            wb = load_workbook(excel_file_path)
            ws = wb["ImageMetadata"]
            ws.append([username, bot_timestamp, filename, extracted_image_timestamp_str])
            wb.save(excel_file_path)
            logging.info(f"✅ Inserted record for '{filename}' into local Excel file: '{excel_file_path}'.")
        except Exception as e:
            logging.error(f"❌ Local Excel write error for '{filename}' to '{excel_file_path}': {e}")


# 
def save_data_to_local_excel_only(username, bot_timestamp, filename, extracted_image_timestamp_str, excel_base_folder):
    """
    Saves data to local Excel file only.
    """
    current_dt = datetime.now()
    
    append_to_local_excel(
        username, bot_timestamp, filename, extracted_image_timestamp_str,
        current_datetime=current_dt,
        base_folder=excel_base_folder
    )