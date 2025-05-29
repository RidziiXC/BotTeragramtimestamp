import logging
import os
import glob
import re
from datetime import datetime
import asyncio
import threading
from openpyxl import load_workbook 
from telegram.ext import ContextTypes 

# --- Import modules from the project ---
import excel_manager # <--- Import excel_manager
import sqlite_manager # <--- Import sqlite_manager


def get_processed_image_filenames_for_resume(excel_base_folder_param): # ลบ google_sheet_id_param, google_sheets_credentials_file_param, google_sheets_scope_param
    """
    Reads all processed image filenames from local Excel files.
    """
    processed_files = set()

    # 1. Read from Local Excel Files
    for user_dir in os.listdir(excel_base_folder_param):
        user_path = os.path.join(excel_base_folder_param, user_dir)
        if os.path.isdir(user_path):
            for excel_file_name in glob.glob(os.path.join(user_path, '*.xlsx')):
                with excel_manager.excel_lock: # ใช้ Lock เมื่ออ่านไฟล์ Excel เพื่อป้องกันการ Conflict
                    try:
                        wb = load_workbook(excel_file_name)
                        ws = wb["ImageMetadata"]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and len(row) > 2:
                                processed_files.add(row[2])
                    except Exception as e:
                        logging.error(f"Error reading processed filenames from local Excel '{excel_file_name}': {e}")
    

    return processed_files

def find_unprocessed_images_for_resume(image_folder_param, processed_files_set):
    """
    Scans the image_folder for image files that are not in the processed_files set.
    Returns a list of full paths to unprocessed images.
    Expects structure: image_folder/username/date/filename.jpg
    """
    unprocessed_images = []
    for root, dirs, files in os.walk(image_folder_param):
        for img_file_name in files:
            if img_file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                if img_file_name not in processed_files_set:
                    full_path = os.path.join(root, img_file_name)
                    unprocessed_images.append(full_path)
    return unprocessed_images

def process_single_unprocessed_image_for_resume(loop, bot_instance, full_image_path: str, 
                                                extract_timestamp_func, insert_missed_record_func, save_data_to_local_excel_func, # เปลี่ยนชื่อ param
                                                excel_base_folder_param): # ลบ google_sheet_id_param, google_sheets_credentials_file_param, google_sheets_scope_param
    """
    Processes a single unprocessed image (saving data to local Excel only).
    """
    logging.info(f"[RESUME] Processing unprocessed image: {full_image_path}")

    filename_with_suffix = os.path.basename(full_image_path)
    username_match = re.match(r'(.+)-log\d{4}-\d{2}-\d{2}-', filename_with_suffix)
    username = username_match.group(1) if username_match else "unknown_user"
    
    bot_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    extracted_image_timestamp_for_excel = bot_timestamp 
    
    try:
        # เรียกใช้ฟังก์ชันบันทึกข้อมูลเฉพาะ Local Excel
        save_data_to_local_excel_func(
            username, bot_timestamp, filename_with_suffix, extracted_image_timestamp_for_excel,
            excel_base_folder=excel_base_folder_param
        )
        logging.info(f"[RESUME] ✅ Finished inserting record for '{filename_with_suffix}'.")
    except Exception as e:
        logging.error(f"[RESUME] ❌ Error saving data for '{filename_with_suffix}': {e}")
        
    logging.info(f"[RESUME] Finished processing for {filename_with_suffix}")

def resume_unprocessed_tasks_init(bot_instance_param, image_folder_param, excel_base_folder_param, # ลบ google_sheet_id_param, google_sheets_credentials_file_param, google_sheets_scope_param
                                  extract_timestamp_func, insert_missed_record_func, save_data_to_local_excel_func): # เปลี่ยนชื่อ param
    """
    Initiates the process of finding and processing any unprocessed images.
    This runs when the bot starts up.
    """
    logging.info("Checking for any unprocessed images from previous sessions...")
    
    processed_files_set = get_processed_image_filenames_for_resume(excel_base_folder_param) # ลบ params
    
    unprocessed_images = find_unprocessed_images_for_resume(image_folder_param, processed_files_set)
    
    if unprocessed_images:
        logging.info(f"Found {len(unprocessed_images)} unprocessed images. Starting background processing...")
        current_loop = asyncio.get_event_loop()
        for img_path in unprocessed_images:
            thread = threading.Thread(target=process_single_unprocessed_image_for_resume,
                                      args=(current_loop, bot_instance_param, img_path,
                                            extract_timestamp_func, insert_missed_record_func, save_data_to_local_excel_func,
                                            excel_base_folder_param)) # ลบ params
            thread.start()
    else:
        logging.info("No unprocessed images found. All tasks are up-to-date.")